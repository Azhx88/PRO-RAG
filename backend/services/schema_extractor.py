import pandas as pd
import openpyxl
import json
from typing import Dict, Any

def extract_excel_schema(file_path: str) -> Dict[str, Any]:
    """
    Extracts schema from Excel file.
    Handles multiple sheets, merged cells, and auto-detects headers.
    Returns JSON-serializable schema dict.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    schema = {"sheets": [], "file_type": "excel"}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Unmerge merged cells — fill with top-left value
        for merged_range in list(ws.merged_cells.ranges):
            min_row, min_col = merged_range.min_row, merged_range.min_col
            top_left_value = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(merged_range))
            for row in ws.iter_rows(
                min_row=min_row, max_row=merged_range.max_row,
                min_col=min_col, max_col=merged_range.max_col
            ):
                for cell in row:
                    cell.value = top_left_value

        # Auto-detect header row: first row where 70%+ cells are non-empty
        header_row_idx = 0
        for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
            non_empty = sum(1 for c in row if c is not None)
            if len(row) > 0 and non_empty / len(row) >= 0.7:
                header_row_idx = i
                break

        # Read with pandas from detected header row
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_idx)
        df.dropna(how="all", inplace=True)

        columns = []
        for col in df.columns:
            sample_vals = df[col].dropna().head(3).tolist()
            col_type = str(df[col].dtype)
            columns.append({
                "name": str(col),
                "dtype": col_type,
                "sample_values": [str(v) for v in sample_vals],
                "null_count": int(df[col].isnull().sum()),
                "row_count": len(df)
            })

        schema["sheets"].append({
            "sheet_name": sheet_name,
            "columns": columns,
            "row_count": len(df),
            "header_row": header_row_idx
        })

    return schema


def extract_csv_schema(file_path: str) -> Dict[str, Any]:
    df = pd.read_csv(file_path)
    columns = []
    for col in df.columns:
        sample_vals = df[col].dropna().head(3).tolist()
        columns.append({
            "name": str(col),
            "dtype": str(df[col].dtype),
            "sample_values": [str(v) for v in sample_vals],
            "null_count": int(df[col].isnull().sum()),
            "row_count": len(df)
        })
    return {
        "file_type": "csv",
        "sheets": [{"sheet_name": "main", "columns": columns, "row_count": len(df)}]
    }
