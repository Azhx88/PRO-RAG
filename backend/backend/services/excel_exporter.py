import os
import uuid
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint


# ── Palette ───────────────────────────────────────────────────────────────────
C_INDIGO      = "3730A3"
C_INDIGO_MID  = "4F46E5"
C_SLATE_DARK  = "1E293B"
C_SLATE_MID   = "334155"
C_SLATE_LIGHT = "F1F5F9"
C_WHITE       = "FFFFFF"
C_BORDER      = "CBD5E1"

PIE_COLORS = ["3730A3","4F46E5","6366F1","818CF8","A5B4FC",
              "0EA5E9","38BDF8","7DD3FC","0284C7","075985"]

# ── Style helpers ─────────────────────────────────────────────────────────────
def _font(size=10, bold=False, color=C_SLATE_DARK, name="Calibri", italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def _border(color=C_BORDER, style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Numeric detection ─────────────────────────────────────────────────────────
_NUMERIC_KW = {
    "sum","total","amount","count","price","revenue","sales","qty","quantity",
    "avg","average","percent","gpa","score","marks","grade","rating","rate",
    "value","profit","loss","cost","salary","income","expense","weight","age",
    "height","distance","duration","hours","days","units","volume","spend",
}

def _col_is_numeric(col_name: str, series: pd.Series) -> bool:
    if any(kw in col_name.lower() for kw in _NUMERIC_KW):
        return True
    try:
        converted = pd.to_numeric(series, errors="coerce")
        return converted.notna().sum() / max(len(series), 1) > 0.7
    except Exception:
        return False

def _coerce_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(series)

def _sanitize_col(name: str) -> str:
    RESERVED = {"sum": "total_value", "count": "record_count",
                "avg": "average_value", "max": "max_value", "min": "min_value"}
    return RESERVED.get(name.strip().lower(), name.strip())

def _detect_chart_type(question: str) -> str:
    q = question.lower()
    if any(w in q for w in ("pie","share","proportion","breakdown","distribution","percentage")):
        return "pie"
    if any(w in q for w in ("trend","over time","monthly","yearly","line","growth")):
        return "line"
    return "bar"


# ── Chart builder ─────────────────────────────────────────────────────────────
def _build_chart(ws_data, df: pd.DataFrame, chart_type: str, title: str):
    n_rows = len(df) + 1

    # Find label col (first text-like) and value col (first numeric)
    label_col_idx = 1
    value_col_idx = None
    found_label   = False

    for i, col in enumerate(df.columns, 1):
        is_num = _col_is_numeric(col, df.iloc[:, i - 1])
        if not is_num and not found_label:
            label_col_idx = i
            found_label   = True
        if is_num and value_col_idx is None:
            value_col_idx = i

    if value_col_idx is None:
        value_col_idx = 2

    data_ref   = Reference(ws_data, min_col=value_col_idx, min_row=1, max_row=n_rows)
    labels_ref = Reference(ws_data, min_col=label_col_idx, min_row=2, max_row=n_rows)

    if chart_type == "pie":
        chart = PieChart()
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        for i, color in enumerate(PIE_COLORS[:len(df)]):
            pt = DataPoint(idx=i)
            pt.graphicalProperties.solidFill = color
            chart.series[0].dPt.append(pt)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showCatName = True
        chart.dataLabels.showSerName = False

    elif chart_type == "line":
        chart = LineChart()
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.series[0].graphicalProperties.line.solidFill = C_INDIGO_MID
        chart.series[0].graphicalProperties.line.width = 25000
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True
        chart.y_axis.title = df.columns[value_col_idx - 1]
        chart.x_axis.title = df.columns[label_col_idx - 1]

    else:  # bar
        chart = BarChart()
        chart.type     = "col"
        chart.grouping = "clustered"
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        chart.series[0].graphicalProperties.solidFill = C_INDIGO_MID
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal     = True
        chart.dataLabels.showSerName = False
        chart.y_axis.title  = df.columns[value_col_idx - 1]
        chart.x_axis.title  = df.columns[label_col_idx - 1]
        chart.y_axis.numFmt = '#,##0.##'

    chart.title  = title[:60]
    chart.style  = 2
    chart.width  = 24
    chart.height = 15
    return chart


# ── Dashboard sheet ───────────────────────────────────────────────────────────
def _build_dashboard(ws, question, sql, insight):

    def section_bar(row, label):
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value     = f"   {label}"
        c.font      = _font(10, bold=True, color=C_WHITE)
        c.fill      = _fill(C_INDIGO)
        c.alignment = _align("left", "center")
        ws.row_dimensions[row].height = 22

    def content_block(start_row, end_row, value, code=False):
        ws.merge_cells(f"A{start_row}:G{end_row}")
        c = ws[f"A{start_row}"]
        c.value     = value
        c.font      = _font(9, name="Courier New") if code else _font(10)
        c.fill      = _fill(C_SLATE_LIGHT)
        c.alignment = _align("left", "top", wrap=True)
        c.border    = _border()
        for r in range(start_row, end_row + 1):
            ws.row_dimensions[r].height = 17

    # Title banner
    ws.merge_cells("A1:G3")
    t = ws["A1"]
    t.value     = "   PG-RAG  ·  Analytics Dashboard"
    t.font      = _font(16, bold=True, color=C_WHITE, name="Calibri")
    t.fill      = _fill(C_SLATE_DARK)
    t.alignment = _align("left", "center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30

    # Subtitle
    ws.merge_cells("A4:G4")
    s = ws["A4"]
    s.value     = "   Auto-generated  ·  Text-to-SQL  ·  Powered by Groq + LLaMA 3.3"
    s.font      = _font(9, italic=True, color="94A3B8")
    s.fill      = _fill(C_SLATE_DARK)
    s.alignment = _align("left", "center")
    ws.row_dimensions[4].height = 16

    ws.row_dimensions[5].height = 6  # spacer

    # Query
    section_bar(6, "📋  QUERY")
    ws.merge_cells("A7:G9")
    q = ws["A7"]
    q.value     = question
    q.font      = _font(11, bold=True)
    q.fill      = _fill(C_WHITE)
    q.alignment = _align("left", "center", wrap=True)
    q.border    = _border(C_INDIGO_MID)
    ws.row_dimensions[7].height = 24
    ws.row_dimensions[8].height = 24
    ws.row_dimensions[9].height = 24

    ws.row_dimensions[10].height = 6  # spacer

    # AI Insight
    section_bar(11, "💡  AI INSIGHT")
    content_block(12, 17, insight)

    ws.row_dimensions[18].height = 6  # spacer

    # SQL
    section_bar(19, "🗄️  SQL QUERY USED")
    content_block(20, 24, sql, code=True)

    ws.row_dimensions[25].height = 6  # spacer

    # Visualization label
    section_bar(26, "📊  VISUALIZATION")

    # Column widths
    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 17


# ── Data sheet ────────────────────────────────────────────────────────────────
def _build_data_sheet(ws, df: pd.DataFrame):
    # Header
    for c_idx, col_name in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=c_idx, value=col_name)
        c.font      = _font(10, bold=True, color=C_WHITE)
        c.fill      = _fill(C_SLATE_MID)
        c.border    = _border()
        c.alignment = _align("center", "center")
    ws.row_dimensions[1].height = 22

    # Data rows
    for r_idx, row_data in enumerate(df.itertuples(index=False), 2):
        row_fill = _fill(C_SLATE_LIGHT) if r_idx % 2 == 0 else _fill(C_WHITE)
        for c_idx, value in enumerate(row_data, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=value)
            c.font   = _font(10)
            c.fill   = row_fill
            c.border = _border()
            col_name = df.columns[c_idx - 1]
            if isinstance(value, (int, float)):
                c.alignment = _align("right")
                if any(k in col_name.lower() for k in
                       ("amount","spend","revenue","price","sales","total","cost","salary","income","profit")):
                    c.number_format = '#,##0.00'
                elif any(k in col_name.lower() for k in
                         ("gpa","rate","percent","avg","average","grade","score","rating")):
                    c.number_format = '0.00'
                else:
                    c.number_format = '#,##0.##'
            else:
                c.alignment = _align("left")
        ws.row_dimensions[r_idx].height = 18

    # Auto column widths
    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"

    # Summary / totals row
    summary_row = len(df) + 2
    ws.row_dimensions[summary_row].height = 20
    for c_idx, col in enumerate(df.columns, 1):
        c = ws.cell(row=summary_row, column=c_idx)
        col_letter = get_column_letter(c_idx)
        if _col_is_numeric(col, df.iloc[:, c_idx - 1]):
            c.value         = f"=SUM({col_letter}2:{col_letter}{len(df)+1})"
            c.number_format = '#,##0.##'
            c.font          = _font(10, bold=True, color=C_WHITE)
            c.fill          = _fill(C_INDIGO)
            c.alignment     = _align("right")
        else:
            c.value     = "TOTAL"
            c.font      = _font(10, bold=True, color=C_WHITE)
            c.fill      = _fill(C_INDIGO)
            c.alignment = _align("left")
        c.border = _border()


from config import settings

class ExportManager:
    def __init__(self):
        self.export_dir = os.path.join(settings.upload_dir, "exports")
        os.makedirs(self.export_dir, exist_ok=True)

    def create_dashboard_excel(
        self,
        question: str,
        sql: str,
        results: list,
        insight: str,
        chart_filename: str = None,
    ) -> str:
        wb = Workbook()
        ws_dash = wb.active
        ws_dash.title = "Dashboard"

        _build_dashboard(ws_dash, question, sql, insight)

        ws_data = wb.create_sheet(title="Data")
        df = pd.DataFrame(results) if results else pd.DataFrame()

        if not df.empty:
            df.rename(columns={c: _sanitize_col(c) for c in df.columns}, inplace=True)
            for col in df.columns:
                if _col_is_numeric(col, df[col]):
                    df[col] = _coerce_numeric(df[col])

            _build_data_sheet(ws_data, df)

            chart_type = _detect_chart_type(question)
            chart = _build_chart(ws_data, df, chart_type, question)
            ws_dash.add_chart(chart, "A27")
        else:
            ws_data["A1"] = "No data returned for this query."

        if chart_filename:
            chart_dir  = "/app/charts" if os.name != "nt" else os.path.join(os.path.dirname(__file__), "charts")
            chart_path = os.path.join(chart_dir, chart_filename)
            if os.path.exists(chart_path):
                ws_img = wb.create_sheet(title="Chart Screenshot")
                img = Image(chart_path)
                img.width  = int(img.width  * 0.75)
                img.height = int(img.height * 0.75)
                ws_img.add_image(img, "A1")

        filename = f"dashboard_{uuid.uuid4().hex[:8]}.xlsx"
        wb.save(os.path.join(self.export_dir, filename))
        return filename

    def create_powerbi_excel(
        self,
        question: str,
        sql: str,
        results: list,
        insight: str,
    ) -> str:
        """
        Power BI optimized export - clean tabular data with no merged cells or styling.
        Creates three simple sheets: QueryResults (main data), QueryMetadata, AIInsight
        """
        df = pd.DataFrame(results) if results else pd.DataFrame()
        
        if df.empty:
            # Handle empty results
            df = pd.DataFrame({"message": ["No data returned for this query."]})
        else:
            # Sanitize column names
            df.rename(columns={c: _sanitize_col(c) for c in df.columns}, inplace=True)
            # Coerce numeric columns
            for col in df.columns:
                if _col_is_numeric(col, df[col]):
                    df[col] = _coerce_numeric(df[col])
        
        # Create simple workbook with pandas ExcelWriter
        filename = f"powerbi_{uuid.uuid4().hex[:8]}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Sheet 1: Main query results (clean table)
            df.to_excel(writer, sheet_name='QueryResults', index=False)
        
        return filename